# services/reports_service.py
"""
Servicio para generar reportes ejecutivos profesionales en Excel y PDF
Con branding MACSA, gráficos integrados y diseño premium
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from datetime import datetime
import io
import logging
import os
import base64
from PIL import Image as PILImage

logger = logging.getLogger(__name__)

# Colores corporativos MACSA
MACSA_BLUE = "2196F3"
MACSA_GOLD = "D4AF37"
MACSA_GRAY = "9E9E9E"
MACSA_DARK_BLUE = "1976D2"
MACSA_GREEN = "4CAF50"
MACSA_RED = "F44336"

# Ruta del logo MACSA
LOGO_PATH = "/opt/callcenter-analytics/backend/src/assets/Icono_macsa.jpg"

class ReportsService:
    """Servicio para generación de reportes ejecutivos profesionales"""

    def generate_excel_report(self, data: dict, report_type: str, date_range: dict = None, chart_image: str = None) -> bytes:
        """
        Genera un reporte ejecutivo profesional en Excel con branding MACSA

        Args:
            data: Diccionario con los datos del reporte
            report_type: Tipo de reporte (general, agents, queues, calls)
            date_range: Diccionario con {'start_date': 'YYYY-MM-DD', 'end_date': 'YYYY-MM-DD'}
            chart_image: Imagen en base64 de los gráficos

        Returns:
            Bytes del archivo Excel generado
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        # Estilos corporativos MACSA
        title_font = Font(name='Calibri', bold=True, size=16, color=MACSA_DARK_BLUE)
        subtitle_font = Font(name='Calibri', size=11, color=MACSA_GRAY)
        header_font = Font(name='Calibri', bold=True, size=12, color="FFFFFF")
        system_name_font = Font(name='Calibri', bold=True, size=14, color=MACSA_BLUE)
        date_font = Font(name='Calibri', size=10, color=MACSA_GRAY)

        header_fill = PatternFill(start_color=MACSA_BLUE, end_color=MACSA_BLUE, fill_type="solid")
        gold_fill = PatternFill(start_color=MACSA_GOLD, end_color=MACSA_GOLD, fill_type="solid")

        thin_border = Border(
            left=Side(style='thin', color=MACSA_GRAY),
            right=Side(style='thin', color=MACSA_GRAY),
            top=Side(style='thin', color=MACSA_GRAY),
            bottom=Side(style='thin', color=MACSA_GRAY)
        )

        # === HEADER PROFESIONAL ===
        current_row = 1

        # Logo MACSA a la izquierda
        if os.path.exists(LOGO_PATH):
            try:
                img = XLImage(LOGO_PATH)
                img.width = 80
                img.height = 80
                ws.add_image(img, f'A{current_row}')
            except Exception as e:
                logger.warning(f"No se pudo cargar el logo: {e}")

        # Nombre del sistema en la parte superior derecha
        ws[f'G{current_row}'] = "Call Center Analytics"
        ws[f'G{current_row}'].font = system_name_font
        ws[f'G{current_row}'].alignment = Alignment(horizontal='right', vertical='top')
        ws.merge_cells(f'G{current_row}:H{current_row}')

        # Fecha y hora de generación justo debajo
        current_row += 1
        ws[f'G{current_row}'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws[f'G{current_row}'].font = date_font
        ws[f'G{current_row}'].alignment = Alignment(horizontal='right', vertical='top')
        ws.merge_cells(f'G{current_row}:H{current_row}')

        # Rango de fechas si está disponible
        if date_range:
            current_row += 1
            ws[f'G{current_row}'] = f"Período: {date_range.get('start_date', '')} al {date_range.get('end_date', '')}"
            ws[f'G{current_row}'].font = date_font
            ws[f'G{current_row}'].alignment = Alignment(horizontal='right', vertical='top')
            ws.merge_cells(f'G{current_row}:H{current_row}')

        current_row = 6  # Dejar espacio después del header

        # Título del reporte
        report_titles = {
            'general': 'Reporte General de Llamadas',
            'agents': 'Reporte de Rendimiento de Agentes',
            'queues': 'Reporte de Colas de Atención',
            'calls': 'Detalle de Llamadas',
            'events': 'Reporte de Eventos del Sistema'
        }
        ws[f'A{current_row}'] = report_titles.get(report_type, f'Reporte de {report_type.upper()}')
        ws[f'A{current_row}'].font = title_font
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

        current_row += 2

        # Línea decorativa
        for col in range(1, 9):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = PatternFill(start_color=MACSA_BLUE, end_color=MACSA_BLUE, fill_type="solid")
        ws.row_dimensions[current_row].height = 3

        current_row += 2

        # Preparar imagen temporal si está disponible
        temp_chart_path = None
        if chart_image:
            try:
                # Decodificar la imagen base64
                image_data = base64.b64decode(chart_image.split(',')[1] if ',' in chart_image else chart_image)

                # Guardar temporalmente
                temp_chart_path = f'/tmp/chart_{datetime.now().timestamp()}.png'
                with open(temp_chart_path, 'wb') as f:
                    f.write(image_data)

                # Añadir al Excel
                chart_img = XLImage(temp_chart_path)
                chart_img.width = 600
                chart_img.height = 300
                ws.add_image(chart_img, f'A{current_row}')

                current_row += 18  # Espacio para la imagen
            except Exception as e:
                logger.warning(f"No se pudo preparar gráfico: {e}")
                temp_chart_path = None

        # Generar contenido según tipo de reporte
        if report_type == "general":
            self._add_general_report_to_excel(ws, data, current_row, header_font, header_fill, thin_border)
        elif report_type == "agents":
            self._add_agents_report_to_excel(ws, data, current_row, header_font, header_fill, thin_border, gold_fill)
        elif report_type == "queues":
            self._add_queues_report_to_excel(ws, data, current_row, header_font, header_fill, thin_border)
        elif report_type == "calls":
            self._add_calls_report_to_excel(ws, data, current_row, header_font, header_fill, thin_border)
        elif report_type == "events":
            self._add_events_report_to_excel(ws, data, current_row, header_font, header_fill, thin_border)

        # Ajustar ancho de columnas
        from openpyxl.cell.cell import MergedCell
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                # Obtener column_letter solo si no es una MergedCell
                if column_letter is None and not isinstance(cell, MergedCell):
                    column_letter = cell.column_letter
                try:
                    if cell.value and not isinstance(cell, MergedCell):
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 4, 50)
                ws.column_dimensions[column_letter].width = max(adjusted_width, 14)

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)

        # Limpiar archivo temporal de gráfico si existe
        if temp_chart_path and os.path.exists(temp_chart_path):
            try:
                os.remove(temp_chart_path)
            except Exception as e:
                logger.warning(f"Error eliminando archivo temporal {temp_chart_path}: {e}")

        output.seek(0)
        return output.getvalue()

    def _add_general_report_to_excel(self, ws, data, start_row, header_font, header_fill, border):
        """Agrega datos del reporte general con diseño ejecutivo"""
        # Encabezado de sección
        ws[f'A{start_row}'] = "📊 RESUMEN EJECUTIVO"
        ws[f'A{start_row}'].font = Font(name='Calibri', bold=True, size=13, color=MACSA_DARK_BLUE)
        ws[f'A{start_row}'].fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        ws.merge_cells(f'A{start_row}:B{start_row}')
        ws.row_dimensions[start_row].height = 25

        start_row += 1

        # Headers
        ws[f'A{start_row}'] = "Métrica"
        ws[f'B{start_row}'] = "Valor"
        for col in ['A', 'B']:
            cell = ws[f'{col}{start_row}']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        start_row += 1

        # Datos con formato condicional
        summary_data = [
            ["Total de Llamadas", data.get('total_calls', 0)],
            ["Llamadas Contestadas", data.get('answered_calls', 0)],
            ["Llamadas Abandonadas", data.get('abandoned_calls', 0)],
            ["Tasa de Respuesta", f"{data.get('answer_rate', 0):.1f}%"],
            ["Tiempo Promedio de Espera", f"{data.get('avg_wait_time', 0):.0f} segundos"],
            ["Duración Promedio", f"{data.get('avg_duration', 0):.0f} segundos"],
        ]

        for row_data in summary_data:
            ws[f'A{start_row}'] = row_data[0]
            ws[f'B{start_row}'] = row_data[1]

            ws[f'A{start_row}'].border = border
            ws[f'B{start_row}'].border = border
            ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'B{start_row}'].alignment = Alignment(horizontal='right', vertical='center')
            ws[f'B{start_row}'].font = Font(name='Calibri', bold=True, size=11)

            # Colorear según métricas
            if "Tasa de Respuesta" in row_data[0]:
                value = float(str(row_data[1]).replace('%', ''))
                if value >= 80:
                    ws[f'B{start_row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    ws[f'B{start_row}'].font = Font(name='Calibri', bold=True, size=11, color=MACSA_GREEN)
                elif value >= 60:
                    ws[f'B{start_row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    ws[f'B{start_row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    ws[f'B{start_row}'].font = Font(name='Calibri', bold=True, size=11, color=MACSA_RED)

            start_row += 1

        # Agregar tabla de resumen diario si está disponible
        if 'daily_summary' in data and data['daily_summary']:
            start_row += 2  # Espacio

            # Encabezado de sección
            ws[f'A{start_row}'] = "📅 RESUMEN DIARIO DETALLADO"
            ws[f'A{start_row}'].font = Font(name='Calibri', bold=True, size=13, color=MACSA_DARK_BLUE)
            ws[f'A{start_row}'].fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            ws.merge_cells(f'A{start_row}:E{start_row}')
            ws.row_dimensions[start_row].height = 25

            start_row += 1

            # Headers
            daily_headers = ["Fecha", "Total", "Contestadas", "Perdidas", "Duración Prom."]
            for idx, header in enumerate(daily_headers):
                col_letter = chr(65 + idx)  # A, B, C, D, E
                cell = ws[f'{col_letter}{start_row}']
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            start_row += 1

            # Datos diarios
            for day in data['daily_summary']:
                ws[f'A{start_row}'] = day.get('date', '')
                ws[f'B{start_row}'] = day.get('total_calls', 0)
                ws[f'C{start_row}'] = day.get('answered_calls', 0)
                ws[f'D{start_row}'] = day.get('missed_calls', 0)
                ws[f'E{start_row}'] = f"{day.get('avg_duration', 0):.0f}s"

                for col in ['A', 'B', 'C', 'D', 'E']:
                    cell = ws[f'{col}{start_row}']
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center' if col != 'A' else 'left', vertical='center')

                # Colorear contestadas y perdidas
                ws[f'C{start_row}'].font = Font(name='Calibri', size=10, color=MACSA_GREEN)
                ws[f'D{start_row}'].font = Font(name='Calibri', size=10, color=MACSA_RED)

                start_row += 1

    def _add_agents_report_to_excel(self, ws, data, start_row, header_font, header_fill, border, gold_fill):
        """Agrega datos de agentes con diseño ejecutivo"""
        # Encabezado de sección
        ws[f'A{start_row}'] = "👥 RENDIMIENTO DE AGENTES"
        ws[f'A{start_row}'].font = Font(name='Calibri', bold=True, size=13, color=MACSA_DARK_BLUE)
        ws[f'A{start_row}'].fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        ws.merge_cells(f'A{start_row}:E{start_row}')
        ws.row_dimensions[start_row].height = 25

        start_row += 1

        # Headers
        headers = ["Agente", "Total Llamadas", "Completadas", "Tiempo Total", "Promedio/Llamada"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        start_row += 1

        # Datos de agentes con ranking visual
        agents_list = data.get('agents', [])
        for idx, agent in enumerate(agents_list):
            row_data = [
                agent.get('agent', ''),
                agent.get('total_calls', 0),
                agent.get('completed_calls', 0),
                agent.get('total_talk_time_formatted', '00:00:00'),
                f"{agent.get('avg_talk_time', 0):.0f}s"
            ]

            # Añadir medalla para top 3
            if idx == 0:
                row_data[0] = f"🥇 {row_data[0]}"
            elif idx == 1:
                row_data[0] = f"🥈 {row_data[0]}"
            elif idx == 2:
                row_data[0] = f"🥉 {row_data[0]}"

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=start_row, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(
                    horizontal='center' if col_idx > 1 else 'left',
                    vertical='center'
                )

                # Resaltar top 3 con gradiente dorado
                if idx < 3:
                    opacity = ["FFF9E6", "FFFBEE", "FFFEF7"][idx]
                    cell.fill = PatternFill(start_color=opacity, end_color=opacity, fill_type="solid")
                    if col_idx == 1:
                        cell.font = Font(name='Calibri', bold=True, size=11)

            start_row += 1

    def _add_queues_report_to_excel(self, ws, data, start_row, header_font, header_fill, border):
        """Agrega datos de colas con indicadores visuales"""
        # Encabezado de sección
        ws[f'A{start_row}'] = "📞 RENDIMIENTO POR COLA"
        ws[f'A{start_row}'].font = Font(name='Calibri', bold=True, size=13, color=MACSA_DARK_BLUE)
        ws[f'A{start_row}'].fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        ws.merge_cells(f'A{start_row}:F{start_row}')
        ws.row_dimensions[start_row].height = 25

        start_row += 1

        # Headers
        headers = ["Cola", "Total", "Contestadas", "Abandonadas", "Tasa Respuesta", "Nivel Servicio"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        start_row += 1

        # Datos de colas con semáforo visual
        for queue in data.get('queues', []):
            row_data = [
                queue.get('queue_name', ''),
                queue.get('total_calls', 0),
                queue.get('answered_calls', 0),
                queue.get('abandoned_calls', 0),
                f"{queue.get('answer_rate', 0):.1f}%",
                f"{queue.get('service_level', 0):.1f}%"
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=start_row, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(
                    horizontal='center' if col_idx > 1 else 'left',
                    vertical='center'
                )

                # Semáforo para tasas de respuesta y servicio
                if col_idx in [5, 6]:
                    try:
                        percentage = float(str(value).replace('%', ''))
                        if percentage >= 80:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(name='Calibri', bold=True, size=11, color=MACSA_GREEN)
                            cell.value = f"✓ {value}"
                        elif percentage >= 60:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            cell.value = f"⚠ {value}"
                        else:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(name='Calibri', bold=True, size=11, color=MACSA_RED)
                            cell.value = f"✗ {value}"
                    except:
                        pass

            start_row += 1

    def _add_calls_report_to_excel(self, ws, data, start_row, header_font, header_fill, border):
        """Agrega resumen diario de llamadas"""
        ws[f'A{start_row}'] = "📅 RESUMEN DIARIO DETALLADO"
        ws[f'A{start_row}'].font = Font(name='Calibri', bold=True, size=13, color=MACSA_DARK_BLUE)
        ws[f'A{start_row}'].fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        ws.merge_cells(f'A{start_row}:E{start_row}')
        ws.row_dimensions[start_row].height = 25

        start_row += 1

        headers = ["Fecha", "Total", "Contestadas", "Perdidas", "Duración Prom."]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        start_row += 1

        for day in data.get('daily_summary', []):
            row_data = [
                day.get('date', ''),
                day.get('total_calls', 0),
                day.get('answered_calls', 0),
                day.get('missed_calls', 0),
                f"{day.get('avg_duration', 0):.0f}s"
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=start_row, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='center' if col_idx != 1 else 'left', vertical='center')
                cell.font = Font(name='Calibri', size=10)

            # Colorear contestadas y perdidas
            ws.cell(row=start_row, column=3).font = Font(name='Calibri', size=10, color=MACSA_GREEN)
            ws.cell(row=start_row, column=4).font = Font(name='Calibri', size=10, color=MACSA_RED)

            start_row += 1

    def _add_events_report_to_excel(self, ws, data, start_row, header_font, header_fill, border):
        """Agrega reporte de eventos del sistema"""
        ws[f'A{start_row}'] = "⚡ EVENTOS DEL SISTEMA"
        ws[f'A{start_row}'].font = Font(name='Calibri', bold=True, size=13, color=MACSA_DARK_BLUE)
        ws[f'A{start_row}'].fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        ws.merge_cells(f'A{start_row}:D{start_row}')
        ws.row_dimensions[start_row].height = 25

        start_row += 1

        headers = ["Evento", "Cantidad", "Porcentaje", "Categoría"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        start_row += 1

        events = data.get('events', [])
        for event in events:
            row_data = [
                event.get('event_description', event.get('event', '')),
                event.get('count', 0),
                f"{event.get('percentage', 0):.1f}%",
                event.get('category', 'General')
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=start_row, column=col_idx)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(
                    horizontal='center' if col_idx > 1 else 'left',
                    vertical='center'
                )

            start_row += 1

    def generate_pdf_report(self, data: dict, report_type: str, date_range: dict = None, chart_image: str = None) -> bytes:
        """
        Genera un reporte ejecutivo profesional en PDF con branding MACSA

        Args:
            data: Diccionario con los datos
            report_type: Tipo de reporte
            date_range: Rango de fechas
            chart_image: Imagen en base64 de los gráficos

        Returns:
            Bytes del PDF generado
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=40
        )
        elements = []
        styles = getSampleStyleSheet()

        # === HEADER EJECUTIVO ===
        # Estilos para el header
        system_name_style = ParagraphStyle(
            'SystemName',
            parent=styles['Normal'],
            fontSize=14,
            textColor=colors.HexColor(f'#{MACSA_BLUE}'),
            alignment=TA_RIGHT,
            fontName='Helvetica-Bold'
        )
        datetime_style = ParagraphStyle(
            'DateTime',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor(f'#{MACSA_GRAY}'),
            alignment=TA_RIGHT
        )

        # Construir texto de la columna derecha
        right_content = []
        right_content.append(Paragraph("Call Center Analytics", system_name_style))
        right_content.append(Spacer(1, 0.05*inch))
        right_content.append(Paragraph(
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
            datetime_style
        ))
        if date_range:
            right_content.append(Paragraph(
                f"Período: {date_range.get('start_date', '')} al {date_range.get('end_date', '')}",
                datetime_style
            ))

        # Crear tabla con logo y texto alineados
        header_table_data = []
        if os.path.exists(LOGO_PATH):
            try:
                logo = Image(LOGO_PATH, width=0.8*inch, height=0.8*inch)
                logo.hAlign = 'LEFT'
                header_table_data.append([logo, right_content])
            except Exception as e:
                logger.warning(f"Error cargando logo: {e}")
                header_table_data.append(['', right_content])
        else:
            header_table_data.append(['', right_content])

        if header_table_data:
            header_table = Table(header_table_data, colWidths=[1.2*inch, 4.8*inch])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ]))
            elements.append(header_table)
            elements.append(Spacer(1, 0.3*inch))

        # Línea separadora
        line_table = Table([['']], colWidths=[6*inch])
        line_table.setStyle(TableStyle([
            ('LINEABOVE', (0, 0), (-1, 0), 3, colors.HexColor(f'#{MACSA_BLUE}')),
        ]))
        elements.append(line_table)
        elements.append(Spacer(1, 0.2*inch))

        # Título del reporte
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor(f'#{MACSA_DARK_BLUE}'),
            spaceAfter=10,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )

        report_titles = {
            'general': 'Reporte General de Llamadas',
            'agents': 'Reporte de Rendimiento de Agentes',
            'queues': 'Reporte de Colas de Atención',
            'calls': 'Detalle de Llamadas',
            'events': 'Reporte de Eventos del Sistema'
        }
        title = Paragraph(report_titles.get(report_type, f'Reporte de {report_type.upper()}'), title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))

        # Preparar imagen temporal si existe
        temp_chart_path = None
        if chart_image:
            try:
                image_data = base64.b64decode(chart_image.split(',')[1] if ',' in chart_image else chart_image)
                temp_chart_path = f'/tmp/chart_{datetime.now().timestamp()}.png'
                with open(temp_chart_path, 'wb') as f:
                    f.write(image_data)

                chart_img = Image(temp_chart_path, width=5.5*inch, height=2.75*inch)
                chart_img.hAlign = 'CENTER'
                elements.append(chart_img)
                elements.append(Spacer(1, 0.3*inch))
            except Exception as e:
                logger.warning(f"Error preparando gráfico para PDF: {e}")
                temp_chart_path = None

        # Añadir contenido según el tipo
        if report_type == "general":
            self._add_general_report_to_pdf(elements, data, styles)
        elif report_type == "agents":
            self._add_agents_report_to_pdf(elements, data, styles)
        elif report_type == "queues":
            self._add_queues_report_to_pdf(elements, data, styles)
        elif report_type == "calls":
            self._add_calls_report_to_pdf(elements, data, styles)
        elif report_type == "events":
            self._add_events_report_to_pdf(elements, data, styles)

        # Footer con marca de agua MACSA
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor(f'#{MACSA_GRAY}'),
            alignment=TA_CENTER
        )
        footer = Paragraph(
            f"<i>Clínica MACSA - Sistema de Análisis de Call Center • Documento confidencial</i>",
            footer_style
        )
        elements.append(Spacer(1, 0.5*inch))
        elements.append(footer)

        # Construir el PDF
        doc.build(elements)

        # Limpiar archivo temporal de gráfico si existe
        if temp_chart_path and os.path.exists(temp_chart_path):
            try:
                os.remove(temp_chart_path)
            except Exception as e:
                logger.warning(f"Error eliminando archivo temporal {temp_chart_path}: {e}")

        buffer.seek(0)
        return buffer.getvalue()

    def _add_general_report_to_pdf(self, elements, data, styles):
        """Agrega resumen general al PDF"""
        section_title = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor(f'#{MACSA_BLUE}'),
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph("📊 RESUMEN EJECUTIVO", section_title))

        table_data = [
            ["Métrica", "Valor"],
            ["Total de Llamadas", str(data.get('total_calls', 0))],
            ["Llamadas Contestadas", str(data.get('answered_calls', 0))],
            ["Llamadas Abandonadas", str(data.get('abandoned_calls', 0))],
            [" Tasa de Respuesta", f"{data.get('answer_rate', 0):.1f}%"],
            ["Tiempo Promedio de Espera", f"{data.get('avg_wait_time', 0):.0f} segundos"],
            ["Duración Promedio", f"{data.get('avg_duration', 0):.0f} segundos"],
        ]

        table = Table(table_data, colWidths=[3.5*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{MACSA_BLUE}')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F5F5F5')),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 11),
            ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(f'#{MACSA_GRAY}')),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor(f'#{MACSA_BLUE}')),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))

        # Agregar tabla de resumen diario si está disponible
        if 'daily_summary' in data and data['daily_summary']:
            elements.append(Paragraph("📅 RESUMEN DIARIO DETALLADO", section_title))

            daily_data = [["Fecha", "Total", "Contestadas", "Perdidas", "Duración Prom."]]

            for day in data['daily_summary']:
                daily_data.append([
                    day.get('date', ''),
                    str(day.get('total_calls', 0)),
                    str(day.get('answered_calls', 0)),
                    str(day.get('missed_calls', 0)),
                    f"{day.get('avg_duration', 0):.0f}s"
                ])

            daily_table = Table(daily_data, colWidths=[1.5*inch, 1*inch, 1.2*inch, 1.2*inch, 1.1*inch])
            daily_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{MACSA_BLUE}')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F5F5F5')),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(f'#{MACSA_GRAY}')),
                ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor(f'#{MACSA_BLUE}')),
                # Alternar colores de filas
                *[('BACKGROUND', (0, i), (-1, i), colors.white) for i in range(2, len(daily_data), 2)]
            ]))
            elements.append(daily_table)
            elements.append(Spacer(1, 0.3*inch))

    def _add_agents_report_to_pdf(self, elements, data, styles):
        """Agrega reporte de agentes al PDF"""
        section_title = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor(f'#{MACSA_BLUE}'),
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph("👥 RENDIMIENTO DE AGENTES", section_title))

        table_data = [["Agente", "Llamadas", "Completadas", "Tiempo Total", "Promedio"]]

        agents_list = data.get('agents', [])[:20]
        for idx, agent in enumerate(agents_list):
            agent_name = agent.get('agent', '')
            if idx == 0:
                agent_name = f"🥇 {agent_name}"
            elif idx == 1:
                agent_name = f"🥈 {agent_name}"
            elif idx == 2:
                agent_name = f"🥉 {agent_name}"

            table_data.append([
                agent_name,
                str(agent.get('total_calls', 0)),
                str(agent.get('completed_calls', 0)),
                agent.get('total_talk_time_formatted', '00:00:00'),
                f"{agent.get('avg_talk_time', 0):.0f}s"
            ])

        table = Table(table_data, colWidths=[1.5*inch, 1*inch, 1.2*inch, 1.3*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{MACSA_BLUE}')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F5F5F5')),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(f'#{MACSA_GRAY}')),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor(f'#{MACSA_BLUE}')),
            ('BACKGROUND', (0, 1), (-1, 3), colors.HexColor('#FFF9E6')),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))

    def _add_queues_report_to_pdf(self, elements, data, styles):
        """Agrega reporte de colas al PDF"""
        section_title = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor(f'#{MACSA_BLUE}'),
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph("📞 RENDIMIENTO POR COLA", section_title))

        table_data = [["Cola", "Total", "Contestadas", "Abandonadas", "Tasa Resp.", "Nivel Serv."]]

        for queue in data.get('queues', []):
            table_data.append([
                queue.get('queue_name', ''),
                str(queue.get('total_calls', 0)),
                str(queue.get('answered_calls', 0)),
                str(queue.get('abandoned_calls', 0)),
                f"{queue.get('answer_rate', 0):.1f}%",
                f"{queue.get('service_level', 0):.1f}%"
            ])

        table = Table(table_data, colWidths=[1.5*inch, 0.9*inch, 1.1*inch, 1.1*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{MACSA_BLUE}')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F5F5F5')),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(f'#{MACSA_GRAY}')),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor(f'#{MACSA_BLUE}')),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))

    def _add_calls_report_to_pdf(self, elements, data, styles):
        """Agrega resumen diario de llamadas al PDF"""
        section_title = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor(f'#{MACSA_BLUE}'),
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph("📅 RESUMEN DIARIO DETALLADO", section_title))

        table_data = [["Fecha", "Total", "Contestadas", "Perdidas", "Duración Prom."]]

        for day in data.get('daily_summary', []):
            table_data.append([
                day.get('date', ''),
                str(day.get('total_calls', 0)),
                str(day.get('answered_calls', 0)),
                str(day.get('missed_calls', 0)),
                f"{day.get('avg_duration', 0):.0f}s"
            ])

        table = Table(table_data, colWidths=[1.5*inch, 1*inch, 1.2*inch, 1.2*inch, 1.1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{MACSA_BLUE}')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F5F5F5')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(f'#{MACSA_GRAY}')),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor(f'#{MACSA_BLUE}')),
            # Alternar colores de filas
            *[('BACKGROUND', (0, i), (-1, i), colors.white) for i in range(2, len(table_data), 2)]
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))

    def _add_events_report_to_pdf(self, elements, data, styles):
        """Agrega reporte de eventos al PDF"""
        section_title = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor(f'#{MACSA_BLUE}'),
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )
        elements.append(Paragraph("⚡ EVENTOS DEL SISTEMA", section_title))

        table_data = [["Evento", "Cantidad", "Porcentaje", "Categoría"]]

        for event in data.get('events', []):
            table_data.append([
                event.get('event_description', event.get('event', '')),
                str(event.get('count', 0)),
                f"{event.get('percentage', 0):.1f}%",
                event.get('category', 'General')
            ])

        table = Table(table_data, colWidths=[2*inch, 1*inch, 1.2*inch, 1.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f'#{MACSA_BLUE}')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F5F5F5')),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(f'#{MACSA_GRAY}')),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor(f'#{MACSA_BLUE}')),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))

# Instancia global
reports_service = ReportsService()
